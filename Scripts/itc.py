#!/usr/bin/env python3
"""
Script: itc.py

This script reads your existing IT Excel workbook, filters out any links already present,
fetches full metadata (title and abstract) for the remaining links in new_it_links.txt, or  hl.txt,
corrects any truncated titles by refetching metadata, and outputs both a CSV and an Excel file
without using pandas.

Requirements:
  - openpyxl
  - requests
  - beautifulsoup4

Usage:
  python itc.py
"""

import re
import requests
import openpyxl
from bs4 import BeautifulSoup
import csv

# Configuration
EXCEL_FILE = 'IT_updated.xlsx'
NEW_LINKS_FILE = 'new_it_links.txt'
OUTPUT_CSV = 'new_core_it_papers.csv'
CORRECTED_EXCEL = 'new_core_it_papers_fixed_titles.xlsx'
SEMANTIC_SCHOLAR_URL = 'https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}?fields=title,abstract'


def load_existing_links():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    try:
        link_col = headers.index('Link') + 1
    except ValueError:
        link_col = 3
    existing = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[link_col-1]
        if url:
            existing.add(url.strip())
    return existing


def load_new_links():
    with open(NEW_LINKS_FILE, 'r', encoding='utf-8') as f:
        return [line.strip() for line in f if line.strip()]


def extract_doi(url):
    m = re.search(r'10\.\d{4,9}/\S+', url)
    return m.group(0) if m else None


def fetch_crossref(doi):
    try:
        r = requests.get(f'https://api.crossref.org/works/{doi}', timeout=10)
        r.raise_for_status()
        data = r.json().get('message', {})
        title = (data.get('title') or [''])[0]
        abstract = data.get('abstract') or ''
        abstract = re.sub(r'<.*?>', '', abstract)
        return title.strip(), abstract.strip()
    except:
        return '', ''


def fetch_semantic_scholar(doi):
    try:
        r = requests.get(SEMANTIC_SCHOLAR_URL.format(doi=doi), timeout=10)
        r.raise_for_status()
        data = r.json()
        return (data.get('title') or '').strip(), (data.get('abstract') or '').strip()
    except:
        return '', ''


def fetch_arxiv(arxiv_id):
    try:
        r = requests.get(f'http://export.arxiv.org/api/query?id_list={arxiv_id}', timeout=10)
        r.raise_for_status()
        from xml.etree import ElementTree as ET
        root = ET.fromstring(r.text)
        ns = {'atom':'http://www.w3.org/2005/Atom'}
        entry = root.find('atom:entry', ns)
        if entry is None:
            return '', ''
        title = entry.find('atom:title', ns).text or ''
        abstract = entry.find('atom:summary', ns).text or ''
        return title.strip(), abstract.strip()
    except:
        return '', ''


def fetch_metadata(url):
    title = abstract = ''
    doi = extract_doi(url)
    if doi:
        title, abstract = fetch_semantic_scholar(doi)
        if not abstract:
            t2, a2 = fetch_crossref(doi)
            if a2:
                title = title or t2
                abstract = a2
    elif 'arxiv.org/abs/' in url:
        title, abstract = fetch_arxiv(url.split('/')[-1])
    if not title or not abstract:
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, 'html.parser')
            if not title:
                meta = soup.find('meta', attrs={'name':'citation_title'}) or soup.find('meta', attrs={'property':'og:title'})
                title = meta['content'].strip() if meta and meta.get('content') else title
            if not abstract:
                meta = soup.find('meta', attrs={'name':'citation_abstract'}) or soup.find('meta', attrs={'property':'og:description'})
                if meta and meta.get('content'):
                    abstract = meta['content'].strip()
                else:
                    div = soup.find('div', class_='abstract') or soup.find('section', class_='abstract')
                    abstract = div.get_text(strip=True) if div else abstract
        except:
            pass
    return title, abstract


def write_to_excel(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Title', 'Abstract', 'Link'])
    for r in rows:
        ws.append([r['Title'], r['Abstract'], r['Link']])
    wb.save(path)


def main():
    existing = load_existing_links()
    candidates = load_new_links()
    to_process = [u for u in candidates if u not in existing]
    print(f"Processing {len(to_process)} new links...")

    rows = []
    for url in to_process:
        title, abstract = fetch_metadata(url)
        if title and abstract:
            rows.append({'Title': title, 'Abstract': abstract, 'Link': url})
        else:
            print(f"Skipped {url}: metadata not found")

    # Correct titles by refetching if truncated
    for r in rows:
        full_title, _ = fetch_metadata(r['Link'])
        if full_title and len(full_title) > len(r['Title']):
            r['Title'] = full_title

    # Write CSV
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['Title', 'Abstract', 'Link'])
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    # Write Excel
    write_to_excel(rows, CORRECTED_EXCEL)
    print(f"Done. Excel file saved to {CORRECTED_EXCEL}")

if __name__ == '__main__':
    main()
    