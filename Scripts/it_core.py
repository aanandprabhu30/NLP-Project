# -*- coding: utf-8 -*-
"""it_core.py

Pandas‑free IT paper harvester with:
  • meta‑tag extraction (works for USENIX/IEEE landing pages)
  • CrossRef JSON fallback for any DOI (bypasses ACM/Elsevier 403)
  • strict core‑IT keyword filter
  • duplicate detection by Title

Usage
-----
1. Put candidate URLs (landing pages **or** https://doi.org/...) in
   `new_it_links.txt`, one per line.
2. Ensure your current IT workbook (e.g. `IT.xlsx`)
   is in the same folder.  Edit BASE_XLSX below if filename differs.
3. Run:
       pip install requests beautifulsoup4 openpyxl
       python fetch_it_core_add_nopandas_v2.py

New rows are appended and saved to `IT_updated.xlsx`.
"""

import json
import re
import sys
import urllib.parse
from pathlib import Path
from typing import List, Tuple

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# ---------------------------- Config -------------------------------------- #
INPUT_FILE  = Path("new_it_links.txt")
BASE_XLSX   = Path("IT_updated.xlsx")   # change if your workbook differs
OUTPUT_XLSX = Path("IT_updated.xlsx")

HEADERS = {"User-Agent": "Mozilla/5.0 (core-it-fetcher/4.0)"}

INCLUDE_KW = [
    "infrastructure", "virtualization", "virtual machine", "vm ",
    "kubernetes", "docker", "container", "orchestration",
    "ansible", "terraform", "infrastructure as code", "iac",
    "aiops", "observability", "monitoring", "sre", "prometheus", "grafana",
    "finops", "cost optim", "billing",
    "edge computing", "edge-cloud", "fog computing",
    "data center", "datacentre", "datacenter",
    "sdn", "nfv", "load balancer", "wan optimization",
    "disaster recovery", "business continuity", "fault tolerance",
    "cloud", "virtualization", "network", "firewall", "container",
    "orchestration", "datacenter", "edge computing", "IaC", "Terraform",
    "DevOps", "AIOps", "microservice", "serverless", "service mesh"
]

EXCLUDE_KW = [
    "adoption", "erp", "information system", "digital transformation",
    "user study", "interface", "education", "healthcare", "blockchain",
    "programming language", "machine learning algorithm", "deep learning",
    "survey of", "human", "social", "economics", "smart contract"
]

# ---------------------------- Helpers ------------------------------------- #
def fetch_landing_or_doi(url: str) -> Tuple[str, str]:
    """Return (html_text, doi). One of them may be empty."""
    if "doi.org/" in url:
        return "", url.split("doi.org/")[1]
    return (requests.get(url, headers=HEADERS, timeout=30).text, "")

def parse_html(html: str) -> Tuple[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    # --- after soup = BeautifulSoup(html, "html.parser") ---

    # 1. Springer / general meta tags
    meta_t = (
        soup.find("meta", {"name": re.compile("citation_title|dc.title", re.I)})
        or soup.find("meta", {"property": "og:title"})
    )
    meta_a = (
        soup.find("meta", {"name": re.compile("citation_abstract|dc.description", re.I)})
    )

    if meta_t and meta_a:
        return meta_t["content"].strip(), meta_a["content"].strip()

    # 2. Springer HTML <section class="Abstract">
    abs_par = soup.select_one("section.Abstract p")
    title_tag = soup.find("h1") or soup.find("title")
    title = title_tag.get_text(" ", strip=True) if title_tag else ""
    abstract = abs_par.get_text(" ", strip=True) if abs_par else ""
    

    # USENIX / IEEE meta tags
    mt = soup.find("meta", {"name": re.compile("citation_title", re.I)})
    ma = soup.find("meta", {"name": re.compile("citation_abstract", re.I)})
    if mt and ma:
        return mt["content"].strip(), ma["content"].strip()

    # USENIX Drupal div
    abs_div = soup.select_one("div.field--name-field-abstract div.field__item")
    title_tag = soup.find("h1") or soup.find("title")
    title = title_tag.get_text(" ", strip=True) if title_tag else ""
    abstract = abs_div.get_text(" ", strip=True) if abs_div else ""

    # Generic fallback
    if not abstract:
        g = soup.find("div", class_=re.compile("abstract", re.I))
        if g:
            abstract = g.get_text(" ", strip=True)

    return title, abstract

def crossref_json(doi: str) -> Tuple[str, str]:
    api = f"https://api.crossref.org/works/{urllib.parse.quote(doi)}"
    try:
        msg = requests.get(api, headers=HEADERS, timeout=25).json()["message"]
        title = msg.get("title", [""])[0]
        abstract = re.sub("<[^>]+>", "", msg.get("abstract", "")).strip()
        return title, abstract
    except Exception as e:
        print(f"[!] CrossRef JSON fail {doi}: {e}", file=sys.stderr)
        return "", ""

def is_core(text: str) -> bool:
    t = text.lower()
    return any(k in t for k in INCLUDE_KW) and not any(k in t for k in EXCLUDE_KW)

def load_titles(xlsx: Path) -> set[str]:
    ws = load_workbook(xlsx, read_only=True).active
    titles = {str(r[0]).strip().lower() for r in ws.iter_rows(min_row=2, values_only=True)}
    ws.parent.close()
    return titles

def append_rows(xlsx: Path, rows: List[Tuple[str, str, str, str]]):
    if xlsx.exists():
        wb, ws = load_workbook(xlsx), None
        ws = wb.active
    else:
        wb, ws = Workbook(), Workbook().active
        ws.append(["Title", "Abstract", "Discipline", "Link"])
    for r in rows:
        ws.append(list(r))
    wb.save(xlsx); wb.close()

def is_arxiv_cs_it(url):
    resp = requests.get(url)
    soup = BeautifulSoup(resp.text, "html.parser")
    primary = soup.find("span", class_="primary-subject")
    return primary and "CS > Information Technology" in primary.text

def extract_meta_arxiv(url: str) -> tuple[str, str]:
    """
    Fetches an arXiv abstract page and returns (title, abstract).
    """
    resp = requests.get(url)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Title is in <h1 class="title mathjax">
    title_el = soup.find("h1", class_="title")
    title = title_el.get_text(strip=True).replace("Title:", "").strip()

    # Abstract is in <blockquote class="abstract mathjax">
    abs_el = soup.find("blockquote", class_="abstract")
    abstract = abs_el.get_text(strip=True).replace("Abstract:", "").strip()

    return title, abstract

# ─────────────────────────── main ───────────────────────────────
def main():
    if not INPUT_FILE.exists(): sys.exit("new_it_links.txt missing")
    if not BASE_XLSX.exists(): sys.exit(f"{BASE_XLSX} missing")

    urls = [u.strip() for u in INPUT_FILE.read_text().splitlines() if u.strip()]
    print(f"Processing {len(urls)} URLs…")
    known = load_titles(BASE_XLSX)
    new_rows = []

    for url in urls:
        print("→", url)
        # Auto-accept **all** arXiv abstracts
        if "arxiv.org/abs" in url:
            title, abstract = extract_meta_arxiv(url)
            new_rows.append((title, abstract, "IT", url))
            known.add(title.lower())
            continue

        # 2️⃣ All other URLs
        html, doi = fetch_landing_or_doi(url)
        title, abstract = "", ""
        if html:
            title, abstract = parse_html(html)
        if not title and doi:
            title, abstract = crossref_json(doi)
        if not title or not abstract:
            print("   • missing title/abstract — skipped")
            continue  # ← this continue matches the outer for-loop

        if title.lower() in known:
            print("   • duplicate — skipped")
            continue

        if not is_core(f"{title} {abstract}"):
            print("   • not core-IT — skipped")
            continue

        new_rows.append((title, abstract, "IT", url))
        known.add(title.lower())
        
        new_rows.append((title, abstract, "IT", url))
        known.add(title.lower())

    if not new_rows:
        print("No new papers added."); return
    if not OUTPUT_XLSX.exists():
        BASE_XLSX.rename(OUTPUT_XLSX)
    append_rows(OUTPUT_XLSX, new_rows)
    print(f"✅ Added {len(new_rows)} papers; IT total ≈ {len(known)}")
    print("Saved →", OUTPUT_XLSX)

if __name__ == "__main__":
    main()