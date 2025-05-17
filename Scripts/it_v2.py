#!/usr/bin/env python3
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

# === Helper functions ===

def extract_meta_arxiv(url: str) -> tuple[str, str]:
    resp = requests.get(url)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    title_el = soup.find("h1", class_="title")
    title = title_el.get_text(strip=True).replace("Title:", "").strip() if title_el else ""
    abs_el = soup.find("blockquote", class_="abstract")
    abstract = abs_el.get_text(strip=True).replace("Abstract:", "").strip() if abs_el else ""
    return title, abstract

def is_arxiv_cs_it(url: str) -> bool:
    resp = requests.get(url)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    prim = soup.find("span", class_="primary-subject")
    return bool(prim and "Information Technology (cs.IT)" in prim.text)

def fetch_landing_or_doi(url: str) -> tuple[str, str]:
    # TODO: Paste your existing implementation here.
    return "", ""

def parse_html(html: str) -> tuple[str, str]:
    # TODO: Paste your existing implementation here.
    return "", ""

def crossref_json(doi: str) -> tuple[str, str]:
    # TODO: Paste your existing implementation here.
    return "", ""

def is_core(text: str) -> bool:
    # TODO: Paste your existing is_core() logic here.
    return True

# === Main processing ===

def main():
    txt_path = "new_it_links.txt"
    excel_path = "IT_updated.xlsx"

    # Load URLs
    if not os.path.exists(txt_path):
        print(f"Link file not found: {txt_path}")
        return
    with open(txt_path) as f:
        urls = [line.strip() for line in f if line.strip()]

    # Prepare workbook
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        # Find Title column index
        headers = [cell.value for cell in ws[1]]
        if "Title" in headers:
            title_idx = headers.index("Title") + 1
        else:
            title_idx = 1
        # Collect known titles
        known = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            known.add(str(row[title_idx-1]).lower())
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Abstract", "Discipline", "Link"])
        known = set()

    new_count = 0

    # Process each URL
    for url in urls:
        print("→", url)

        if "arxiv.org/abs" in url:
            ok = is_arxiv_cs_it(url)
            print(f"   • {url} {'✔ cs.IT' if ok else '✘ NOT cs.IT'}")
            title, abstract = extract_meta_arxiv(url)
            if title and abstract and title.lower() not in known:
                ws.append([title, abstract, "IT", url])
                known.add(title.lower())
                new_count += 1
            continue

        html, doi = fetch_landing_or_doi(url)
        title, abstract = "", ""
        if html:
            title, abstract = parse_html(html)
        if not title and doi:
            title, abstract = crossref_json(doi)
        if not title or not abstract:
            print("   • missing title/abstract — skipped")
            continue
        if title.lower() in known:
            print("   • duplicate — skipped")
            continue
        if not is_core(f"{title} {abstract}"):
            print("   • not core-IT — skipped")
            continue
        ws.append([title, abstract, "IT", url])
        known.add(title.lower())
        new_count += 1

    # Save workbook
    wb.save(excel_path)
    print(f"Added {new_count} papers; saved → {excel_path}")

if __name__ == "__main__":
    main()
